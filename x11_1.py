#!/usr/bin/env python3

#copy version.
import collections
import openpyxl.utils.cell
import g_var
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import Alignment
from getcoor import *
from check_unique import *
from get_var import *
try:
  from openpyxl.cell import get_column_letter
except ImportError:
  print("importError")
  from openpyxl.utils import get_column_letter
  from openpyxl.utils import column_index_from_string

wb = Workbook()
#rawfile = '/Users/jo/desktop/raw_1.xlsx' #origin
#reffile = '/Users/jo/desktop/rf.xlsx' #origin
reffile = get_sel_ref_path()
rawfile = get_sel_raw_path()


raw = load_workbook(rawfile)
ref = load_workbook(reffile)

raw_sheet = raw.active
ref_sheet = ref.active
ws = wb.create_sheet("ISIR_List", 0)
ws1 = wb.create_sheet("Fault_log", 1)

project_code = get_pj_code() 
raw_coor = ''

raw_max_col = raw_sheet.max_column
raw_max_col_lt = openpyxl.utils.cell.get_column_letter(raw_max_col)

#geting part name coordinate in raw sheet.
for cell in ref_sheet[1:1]:
#  for cell in rows:
  if str(cell.value).upper() == 'PART NAME':
    print(str(cell.value).upper())  #expect : PART NAME
    print(cell.column_letter) #expect : N
    ref_coor = cell.column_letter+str((cell.row)+1) 
    ref_pname_start_row = (cell.row)+1
    ref_pname_start_col = cell.column
    # get part name value start coordinate < end
print(ref_coor) # expect : N8

#getting coordinate by getcoor module

res = {}  #result dictionary assign.

#getting a max row line in reference both sheets. for looping limit. 
ref_max_row = ref_sheet.max_row
raw_max_row = raw_sheet.max_row

#supplier column index. 
sup_col = get_sup_col_ltr(raw_sheet)

#finding matched part name in raw sheet.
for rows in raw_sheet['A6':raw_max_col_lt+'7']:
  for cell in rows:
    if str(cell.value).upper() == 'PART NAME':
      print(cell)
      raw_coor = cell.column_letter+str((cell.row)+1)
      raw_pname_start_row = (cell.row)+2 # add 2 cause merge cell.
      raw_pname_start_col = cell.column
      
print(raw_coor)

log_cnt = 2
#need to change cell coordinate range! cause ref_coor coulumn letter not one character.
#using iter_cols
#for ref_cells in ref_sheet[ref_coor:ref_coor[:1]+str(ref_max_row)]:
for ref_cells in ref_sheet.iter_cols(min_row=ref_pname_start_row, min_col=ref_pname_start_col, max_row=ref_max_row, max_col=ref_pname_start_col):
  for ref_cell in ref_cells:
    find_flag = False
    #for raw_cells in raw_sheet[raw_coor:raw_coor[:1]+str(raw_max_row)]:
    for raw_cells in raw_sheet.iter_cols(min_row=raw_pname_start_row, min_col=raw_pname_start_col, max_row=raw_max_row, max_col=raw_pname_start_col):
      for raw_cell in raw_cells:
        if ref_cell.value == raw_cell.value:
          find_flag = True
          res[raw_sheet[get_pno_col_ltr(raw_sheet)+str(raw_cell.row)].value] = {
            'upg': raw_sheet[get_upg_col_ltr(raw_sheet)+str(raw_cell.row)].value,
            'part name': raw_sheet[get_pname_col_ltr(raw_sheet)+str(raw_cell.row)].value,
            'charge': ref_sheet[get_charge_col_ltr(ref_sheet)+str(ref_cell.row)].value,
            'eo no.': raw_sheet[get_eono_col_ltr(raw_sheet)+str(raw_cell.row)].value,
         #   'eo date': raw_sheet[get_eodate_col_ltr(raw_sheet)+str(raw_cell.row)].value,
            'supplier': get_sup_name(raw_sheet, sup_col, raw_cell.row)
            }
    if find_flag == False:
      ws1['B'+str(log_cnt)] = "item not found :: "+ref_cell.value
      log_cnt += 1
# sort dictionary by key using collection module specific OrderedDict()
sor_res = collections.OrderedDict(sorted(res.items()))

print(sor_res)

#making dictionary end.

#ws add dictionary.
row = 3
for key, values in sor_res.items():
#  for v_key, value in res[key].items():
  ws['D'+str(row)] = values['upg']
  ws['E'+str(row)] = key
  ws['F'+str(row)] = values['part name']
  ws['H'+str(row)] = values['charge']
  ws['K'+str(row)] = values['eo no.']
  ws['I'+str(row)] = values['supplier']
  ws['A'+str(row)] = row-2
  row += 1

#Alignment 
get_max_row = ws.max_row
for rows in ws['A3':'E'+str(get_max_row)]:
  for cell in rows:
    cell.alignment = Alignment(horizontal='center', vertical='center') 

for rows in ws['F3':'F'+str(get_max_row)]:
  for cell in rows: 
    cell.alignment = Alignment(vertical='center')  

for rows in ws['G3':'M'+str(get_max_row)]:
  for cell in rows:
    cell.alignment = Alignment(horizontal='center', vertical='center')

#checking asterisk
for rows in ws['F3':'F'+str(get_max_row)]:
  for cell in rows:
    if cell.row == 3:
      if check_unique(ws, cell.row, project_code) == True:
        ws['B'+str(cell.row)] = '*'
        ws['C'+str(cell.row)] = '*'
      else:
        ws['B'+str(cell.row)] = '.'
        ws['C'+str(cell.row)] = '.'
    else:
      if check_unique(ws, cell.row, project_code) == True:
        if cell.value == ws['F'+str((cell.row)-1)].value:
          ws['C'+str(cell.row)] = '*'
        else:
          ws['B'+str(cell.row)] = '*'
          ws['C'+str(cell.row)] = '*'
      else:
        if cell.value == ws['F'+str((cell.row)-1)].value:
          ws['C'+str(cell.row)] = '.'
        else:
          ws['B'+str(cell.row)] = '.'
          ws['C'+str(cell.row)] = '.'

# styles.
titles = ['No.', 'ITEM', 'PARTS','UPG', 'PART-NO', 'PART NAME', 'IRE', 'CHARGE', 'SUPPLIER', 'DESCRIPTION', 'EO-NO', 'R&D', 'PARTS DEVELOP.']
fill = styles.PatternFill(start_color='202020', end_color='202020', fill_type='solid')
sub_font = styles.Font(bold=True, size=24, color='000000FF')
border1 = styles.Border(bottom=styles.Side(border_style='thick', color='00C0C0C0'))
border2 = styles.Border(left=styles.Side(border_style='thin',
                                        color='00C0C0C0'),
                        right=styles.Side(border_style='thin',
                                        color='00C0C0C0'),
                        top=styles.Side(border_style='thin',
                                        color='00C0C0C0'),
                        bottom=styles.Side(border_style='thin',
                                        color='00C0C0C0')
                       )
t_cnt = 1
for title in titles:
  ws.cell(row=2, column=t_cnt, value=title)
  t_cnt += 1

for row in ws['A2':'M2']:
  for cell in row:
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = styles.Font(color='B2FF66', bold=True)
    cell.fill = fill
    cell.border = border2

for column_cells in ws.columns:
  new_column_length = max(len(str(cell.value)) for cell in column_cells)
  new_column_letter = (get_column_letter(column_cells[0].column))
  if new_column_length > 0:
    ws.column_dimensions[new_column_letter].width = new_column_length*1.3

#sub_font = styles.Font(bold=True, size=24, color='000000FF')
#border1 = styles.Border(bottom=styles.Side(border_style='thick'))
#border2 = styles.Border(left=styles.Side(border_style='thin',
#                                        color='00C0C0C0'),
#                        right=styles.Side(border_style='thin',
#                                        color='00C0C0C0'),
#                        top=styles.Side(border_style='thin',
#                                        color='00C0C0C0'),
#                        bottom=styles.Side(border_style='thin',
#                                        color='00C0C0C0')
#                       )
#fill = styles.PatternFill(start_color='E2E2EF', end_color='E2E2EF', fill_type='solid')


ws['A1'] = project_code+" ISIR PARTS LIST"
ws['A1'].font = sub_font
ws['A1'].border = border1 
ws['A1'].alignment = Alignment(vertical='center')
ws.row_dimensions[1].height = 38 
ws.row_dimensions[2].height = 28 
ws.column_dimensions['L'].width = 15 
ws.column_dimensions['M'].width = 15
ws.column_dimensions['B'].width = 6
ws.column_dimensions['C'].width = 6 

ws.merge_cells('A1:M1')

ws_cell_range = ws['A3':'M'+str(get_max_row)]
for rows in ws_cell_range:
  for cell in rows:
    cell.border = border2
    cell.font = styles.Font(size=10)

#title_range = ws['A2':'M2']
#for title in title_range:
#  for cell in title:
#    cell.fill = fill
##fn = '/Users/jo/desktop/result_2.xlsx' #origin
#fn = get_save_path()
if len(g_var.file_name[0])>1:
  fn = get_save_path()+"/"+g_var.file_name[0]+".xlsx"
  print(g_var.file_name)
else:
  fn = get_save_path()+"/"+project_code+"_"+"Parts LIST.xlsx"
  print(g_var.file_name)
print(get_sel_ref_path())
print(get_sel_raw_path())
print(get_save_path())
print(get_pj_code())
print(fn)

#wb.save(fn)
